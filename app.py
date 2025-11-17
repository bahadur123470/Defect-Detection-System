"""
Main Flask application for Defect Detection System
"""
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
import json
import csv
from datetime import datetime
from io import BytesIO, StringIO
from image_processor import ImageProcessor
from defect_detector import DefectDetector
from report_generator import ReportGenerator
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('reports', exist_ok=True)
os.makedirs('static/processed', exist_ok=True)

# Simple user database (in production, use a real database)
# Using pbkdf2:sha256 method for compatibility with Python 3.9
USERS = {
    'admin': generate_password_hash('admin123', method='pbkdf2:sha256'),
    'user': generate_password_hash('user123', method='pbkdf2:sha256')
}

# Initialize components
image_processor = ImageProcessor()
defect_detector = DefectDetector()
report_generator = ReportGenerator()


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


@app.route('/')
def index():
    """Home page - redirect to login if not authenticated"""
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username in USERS and check_password_hash(USERS[username], password):
            session['username'] = username
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    """User logout"""
    session.pop('username', None)
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))


@app.route('/register', methods=['POST'])
def register():
    """User registration"""
    username = request.form.get('username')
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')

    if not username or not password or not confirm_password:
        flash('All fields are required', 'error')
        return redirect(url_for('login'))

    if password != confirm_password:
        flash('Passwords do not match', 'error')
        return redirect(url_for('login'))

    if username in USERS:
        flash('Username already exists', 'error')
        return redirect(url_for('login'))

    # Hash the password and add to users
    USERS[username] = generate_password_hash(password, method='pbkdf2:sha256')
    flash('Registration successful! Please log in.', 'success')
    return redirect(url_for('login'))


@app.route('/dashboard')
def dashboard():
    """Main dashboard"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # Get list of processed images
    processed_images = []
    processed_dir = 'static/processed'
    if os.path.exists(processed_dir):
        processed_images = [f for f in os.listdir(processed_dir) 
                           if f.endswith(('.png', '.jpg', '.jpeg'))]
    
    return render_template('dashboard.html', 
                         username=session['username'],
                         processed_images=processed_images)


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    """Image upload and processing"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Process image
            try:
                # Preprocess image
                processed_image = image_processor.preprocess(filepath)
                
                # Detect defects
                detection_results = defect_detector.detect(processed_image)
                
                # Save processed image with annotations
                output_filename = f"processed_{filename}"
                output_path = os.path.join('static/processed', output_filename)
                annotated_image = defect_detector.draw_detections(processed_image, detection_results)
                image_processor.save_image(annotated_image, output_path)
                
                # Save detection results
                results_data = {
                    'original_filename': filename,
                    'processed_filename': output_filename,
                    'detections': detection_results,
                    'timestamp': timestamp,
                    'defect_count': len(detection_results)
                }
                
                results_file = os.path.join('static/processed', f"results_{filename}.json")
                with open(results_file, 'w') as f:
                    json.dump(results_data, f, indent=2)
                
                flash(f'Image processed successfully! Found {len(detection_results)} defects.', 'success')
                return redirect(url_for('results', filename=output_filename))
            
            except Exception as e:
                flash(f'Error processing image: {str(e)}', 'error')
                return redirect(request.url)
    
    return render_template('upload.html')


@app.route('/results/<filename>')
def results(filename):
    """Display detection results"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # Load detection results
    results_file = os.path.join('static/processed', f"results_{filename.replace('processed_', '')}.json")
    if os.path.exists(results_file):
        with open(results_file, 'r') as f:
            results_data = json.load(f)
    else:
        results_data = {'detections': [], 'defect_count': 0}
    
    # Compute defect type breakdown
    defect_types = {}
    for det in results_data.get('detections', []):
        defect_type = det.get('type', 'Unknown')
        defect_types[defect_type] = defect_types.get(defect_type, 0) + 1
    results_data['defect_types'] = defect_types if defect_types else {}
    
    return render_template('results.html', 
                         filename=filename,
                         results=results_data)


@app.route('/generate_report/<filename>')
def generate_report(filename):
    """Generate and download report"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Load detection results
        results_file = os.path.join('static/processed', f"results_{filename.replace('processed_', '')}.json")
        if not os.path.exists(results_file):
            flash('Results file not found', 'error')
            return redirect(url_for('dashboard'))
        
        with open(results_file, 'r') as f:
            results_data = json.load(f)
        
        # Generate report
        report_path = report_generator.generate(
            results_data,
            os.path.join('static/processed', filename)
        )
        
        return send_file(report_path, as_attachment=True, 
                        download_name=f"defect_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'error')
        return redirect(url_for('dashboard'))


@app.route('/export_csv/<filename>')
def export_csv(filename):
    """Export detection results as CSV"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Load detection results
        results_file = os.path.join('static/processed', f"results_{filename.replace('processed_', '')}.json")
        if not os.path.exists(results_file):
            flash('Results file not found', 'error')
            return redirect(url_for('dashboard'))
        
        with open(results_file, 'r') as f:
            results_data = json.load(f)
        
        # Create CSV in memory using StringIO
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Defect ID', 'Type', 'Confidence', 'X1', 'Y1', 'X2', 'Y2', 'Area (px²)'])
        
        # Write data
        for idx, det in enumerate(results_data.get('detections', []), 1):
            bbox = det.get('bbox', [0, 0, 0, 0])
            area = (bbox[2] - bbox[0]) * (bbox[3] - bbox[1]) if len(bbox) == 4 else 0
            writer.writerow([
                idx,
                det.get('type', 'Unknown'),
                f"{det.get('confidence', 0):.2f}",
                bbox[0] if len(bbox) > 0 else 0,
                bbox[1] if len(bbox) > 1 else 0,
                bbox[2] if len(bbox) > 2 else 0,
                bbox[3] if len(bbox) > 3 else 0,
                area
            ])
        
        # Convert StringIO to BytesIO for send_file
        output.seek(0)
        csv_bytes = BytesIO(output.getvalue().encode('utf-8'))
        csv_bytes.seek(0)
        
        return send_file(
            csv_bytes,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"defect_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
    
    except Exception as e:
        flash(f'Error exporting CSV: {str(e)}', 'error')
        return redirect(url_for('dashboard'))


@app.route('/export_excel/<filename>')
def export_excel(filename):
    """Export detection results as Excel"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if not OPENPYXL_AVAILABLE:
        flash('Excel export requires openpyxl. Please install it: pip install openpyxl', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Load detection results
        results_file = os.path.join('static/processed', f"results_{filename.replace('processed_', '')}.json")
        if not os.path.exists(results_file):
            flash('Results file not found', 'error')
            return redirect(url_for('dashboard'))
        
        with open(results_file, 'r') as f:
            results_data = json.load(f)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Defect Detection Results"
        
        # Header styling
        header_fill = PatternFill(start_color="FF4D4F", end_color="FF4D4F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        headers = ['Defect ID', 'Type', 'Confidence', 'X1', 'Y1', 'X2', 'Y2', 'Area (px²)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for idx, det in enumerate(results_data.get('detections', []), 1):
            bbox = det.get('bbox', [0, 0, 0, 0])
            area = (bbox[2] - bbox[0]) * (bbox[3] - bbox[1]) if len(bbox) == 4 else 0
            row = [
                idx,
                det.get('type', 'Unknown'),
                round(det.get('confidence', 0), 2),
                bbox[0] if len(bbox) > 0 else 0,
                bbox[1] if len(bbox) > 1 else 0,
                bbox[2] if len(bbox) > 2 else 0,
                bbox[3] if len(bbox) > 3 else 0,
                area
            ]
            for col, value in enumerate(row, 1):
                ws.cell(row=idx + 1, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"defect_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    
    except Exception as e:
        flash(f'Error exporting Excel: {str(e)}', 'error')
        return redirect(url_for('dashboard'))


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)

